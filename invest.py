from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from datetime import datetime
from pytz import timezone
from openapi_client import openapi


try:
    with open('token.txt', 'r') as f:
        data = f.read()
        f.close()
except:
    ok = input('Файл не найден! Создайте файл и введите 1: ')
    if ok == 1 or '1':
        with open('token.txt', 'r') as f:
            data = f.read()
            f.close()
    else:
        print('Вы так ничего и не ввели')

token = data
client = openapi.api_client(token)

class get_instrument():

    @staticmethod
    def price_dollar():
        pf = client.portfolio.portfolio_get()
        baks = pf.payload.positions
        for i in baks:
            if i.name == 'Доллар США':
                balance = i.average_position_price.value
                WM = i.expected_yield.value
                how_much = i.balance
                cb = (balance * how_much) + WM
                weidth = round(cb / how_much, 3)
                return weidth

    @staticmethod
    def all():
        pf = client.portfolio.portfolio_get()
        pf2 = client.portfolio.portfolio_currencies_get()
        position = pf.payload.positions
        currencies = pf2.payload.currencies
        for c in currencies:
            if c.currency == 'RUB':
                free_rub = c.balance
        value, currency, balance, ticker, name, amount_all, current_amount, current_amount_all, instrument_type, amount_all_inst = [], [], [], [], [], [], [], [], [], []
        for i in position:
            value.append(i.average_position_price.value)
            currency.append(i.average_position_price.currency)
            balance.append(i.balance)
            ticker.append(i.ticker)
            name.append(i.name)
            aa = round(i.average_position_price.value * i.balance, 2)
            ca = round((i.average_position_price.value * i.balance + i.expected_yield.value) / i.balance, 2)
            caa = round(i.average_position_price.value * i.balance + i.expected_yield.value, 2)
            amount_all.append(aa)
            current_amount.append(ca)
            current_amount_all.append(caa)
            if i.average_position_price.currency == 'USD':
                value_baks = get_instrument.price_dollar()
                value_in_rub = round(caa * value_baks, 2)
                amount_all_inst.append(value_in_rub)
            else:
                amount_all_inst.append(caa)
            if i.instrument_type == 'Stock':
                instrument_type.append('Акции')
            elif i.instrument_type == 'Bond':
                instrument_type.append('Облигации')
            elif i.instrument_type == 'Etf':
                instrument_type.append('ETF')
            elif i.instrument_type == 'Currency':
                instrument_type.append('Валюта')
            else:
                instrument_type.append(' - ')
        aai = sum(amount_all_inst)
        return value, currency, balance, ticker, name, amount_all, current_amount, current_amount_all, instrument_type, aai, free_rub

class get_operation():
    def dividend(month=1, year=2020, day=1, hour=0, minute=0, second=1, zone='Europe/Moscow'):
        d1 = datetime(year, month, day, hour, minute, second,
                      tzinfo=timezone(zone))
        d2 = datetime.now(tz=timezone(zone))
        ops = client.operations.operations_get(_from=d1.isoformat(), to=d2.isoformat())
        list_inst = get_instrument.all()
        list_name = {}
        for i, k in zip(list_inst[4], list_inst[8]):
            if k == 'Облигации':
                list_name.update({i: [{'Coupon': [], 'TaxCoupon': []}]})
            elif k == 'Акции':
                list_name.update({i: [{'Dividend': [], 'TaxDividend': []}]})
            else:
                pass
            lll = str(i) + 'div_list'
            m = globals()[lll] = []
            kkk = str(i) + 'taxDiv_list'
            t = globals()[kkk] = []
        for op in ops.payload.operations:  # Перебираем операции
            if op.operation_type == 'Dividend':
                a = client.market.market_search_by_figi_get(op.figi)
                name = a.payload.name
                get_operation.list_update(list_name, name, op)
            elif op.operation_type == 'Coupon':
                a = client.market.market_search_by_figi_get(op.figi)
                name = a.payload.name
                get_operation.list_update(list_name, name, op)
            elif op.operation_type == 'TaxCoupon':
                a = client.market.market_search_by_figi_get(op.figi)
                name = a.payload.name
                get_operation.list_update(list_name, name, op)
            elif op.operation_type == 'TaxDividend':
                a = client.market.market_search_by_figi_get(op.figi)
                name = a.payload.name
                get_operation.list_update(list_name, name, op)
            else:
                continue
        list_company_del, list_op_del = [], []
        for i in list_name:
            for m in list_name[i][0]:
                if not list_name[i][0][m]:
                    list_company_del.append(i)
                    list_op_del.append(m)
        for l, h in zip(list_company_del, list_op_del):
            del list_name[l][0][h]
        for a in list(list_name.keys()):
            if not list_name[a][0]:
                del list_name[a]
        result = get_operation.subtraction_taxDiv(list_name)
        return result

    @staticmethod
    def list_update(list_name, name, op, price_in_rub=0):
        if op.currency == 'USD':
            try:
                list_name[name][0][op.operation_type].append(
                    [op.date.strftime('%d.%m.%Y'), round(op.payment, 1), op.currency, round(int(price_in_rub), 1)])
            except:
                list_name.update({name: [{op.operation_type: [
                    [op.date.strftime('%d.%m.%Y'), round(op.payment, 1), op.currency, round(int(price_in_rub), 1)]]}]})
        else:
            try:
                list_name[name][0][op.operation_type].append(
                    [op.date.strftime('%d.%m.%Y'), round(op.payment, 1), op.currency])
            except:
                list_name.update(
                    {name: [{op.operation_type: [[op.date.strftime('%d.%m.%Y'), round(op.payment, 1), op.currency]]}]})

    @staticmethod
    def subtraction_taxDiv(list):
        lis = {}
        for i in list:
            divident, divDate, currency, taxDividenet = [], [], [], []
            for m in list[i][0]:
                if list[i][0][m][0][2] != 'USD':
                    if m == 'Dividend':
                        for l in list[i][0][m]:
                            divident.append(l[1])
                            divDate.append(l[0])
                            currency.append(l[2])
                    elif m == 'Coupon':
                        for l in list[i][0][m]:
                            divident.append(l[1])
                            divDate.append(l[0])
                            currency.append(l[2])
                    else:
                        for l in list[i][0][m]:
                            taxDividenet.append(l[1])
                else:
                    for t in list[i][0][m]:
                        divident.append(t[1])
                        divDate.append(t[0])
                        currency.append(t[2])
                    kek = []
                    for x, y, z in zip(divident, divDate, currency):
                        if len(list[i][0][m]) == 1:
                            kek.append(x)
                            kek.append(y)
                            kek.append(z)
                        else:
                            kek.append([x, y, z])
            if taxDividenet:
                    result = [x+y for x, y in zip(divident, taxDividenet)]
                    a = []
                    if len(result) > 1:
                        g = 0
                        for n in result:
                            kek = []
                            kek.append(n)
                            kek.append(list[i][0][m][g][0])
                            kek.append(list[i][0][m][g][2])
                            a.append(kek)
                            g += 1
                        lis.update({i: a})
                    else:
                        result.append(l[0])
                        result.append(l[2])
                        lis.update({i: result})
            else:
                lis.update({i: kek})
        return lis


def list_alphabet():
    data = [i for i in range(1, 26)]
    alphabet = []
    for i in range(65, 91):
        alphabet.append(chr(i))
    la = {x: y for x,y in zip(data,alphabet)}
    return la


def set_table_1(list, col, row, title):
    ws.cell(column=col, row=row).value = title
    ws.cell(column=col, row=row).border = medium_border
    ws.cell(column=col, row=row).font = Font(size=12, bold=True)
    row+=1
    for i in list:
        ws.cell(column=col, row=row).value = i
        ws.cell(column=col, row=row).border = medium_border
        row+=1


def set_table_2(list):
    ws.column_dimensions['A'].width = 25
    set_table_1(list[4], 1, 1, 'Название')
    set_table_1(list[8], 2, 1, 'Тип')
    set_table_1(list[3], 3, 1, 'Тикер')
    set_table_1(list[2], 4, 1, 'КА')
    set_table_1(list[0], 5, 1, 'СЦП')
    set_table_1(list[6], 6, 1, 'ТС')
    set_table_1(list[7], 7, 1, 'ТОС')
    set_table_1(list[1], 8, 1, 'Валюта')
    lenght = str(len(list[4])+8)
    ws.merge_cells(str('A'+lenght+':G'+lenght))
    ws.merge_cells(str('A'+str(int(lenght)+1)+':G'+str(int(lenght)+1)))
    ws.merge_cells(str('A' + str(int(lenght) + 2) + ':G' + str(int(lenght) + 2)))
    ws.merge_cells(str('A' + str(int(lenght) + 3) + ':G' + str(int(lenght) + 3)))
    ws.merge_cells(str('A' + str(int(lenght) + 4) + ':G' + str(int(lenght) + 4)))
### Оформляется итоговая таблица
    ws.cell(column=1, row=int(lenght) - 5).value = 'Общая стоимость в ₽'
    ws.cell(column=1, row=int(lenght) - 5).border = medium_border
    ws.cell(column=1, row=int(lenght) - 5).font = Font(size=12, bold=True)
    ws.cell(column=2, row=int(lenght) - 5).value = list[9]
    ws.cell(column=2, row=int(lenght) - 5).border = medium_border
    ws.cell(column=1, row=int(lenght) - 4).value = 'Свободные ₽'
    ws.cell(column=1, row=int(lenght) - 4).border = medium_border
    ws.cell(column=1, row=int(lenght) - 4).font = Font(size=12, bold=True)
    ws.cell(column=2, row=int(lenght) - 4).value = list[10]
    ws.cell(column=2, row=int(lenght) - 4).border = medium_border
    ws.cell(column=1, row=int(lenght) - 3).value = 'Итого'
    ws.cell(column=1, row=int(lenght) - 3).border = medium_border
    ws.cell(column=1, row=int(lenght) - 3).font = Font(size=12, bold=True)
    ws.cell(column=2, row=int(lenght) - 3).value = list[9]+list[10]
    ws.cell(column=2, row=int(lenght) - 3).border = medium_border
###
    ws.cell(column=1, row=int(lenght)).value = 'КА - Колличество Акций'
    ws.cell(column=1, row=int(lenght)+1).value = 'СЦП - Средняя цена акции при покупке'
    ws.cell(column=1, row=int(lenght)+2).value = 'ТС - Текущая стоимость'
    ws.cell(column=1, row=int(lenght) + 3).value = 'ТОС - Текущая общая стоимость'


wb = Workbook()
ws = wb.create_sheet('первая страница', 0)

name_fail = (str(datetime.now(tz=timezone('Europe/Moscow')).strftime("%d-%m-%Y"))+'.xlsx')

x = get_instrument.all()
v = get_operation.dividend()

col = 11
title_row = 1
k = 1

medium_border = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))

thin_border = Border(top=Side(style='medium'),
                     bottom=Side(style='medium'))

left_border = Border(left=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))

right_border = Border(right=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))

bukva = list_alphabet()

set_table_2(x)
ws.cell(column=col, row=title_row).value = 'price'
ws.cell(column=col, row=title_row).border = medium_border
ws.cell(column=col + 1, row=title_row).value = 'date'
ws.cell(column=col + 1, row=title_row).border = medium_border
ws.cell(column=col + 2, row=title_row).value = 'currency'
ws.cell(column=col + 2, row=title_row).border = medium_border
ws.column_dimensions[bukva[col-1]].width = 20

for i in v:
    try:
        row = (x[4].index(i))+2
    except:
        row = len(x[4])+2
        k+=1
    ws.cell(column=col - 1, row=row).value=i
    ws.cell(column=col - 1, row=row).border = medium_border
    ws.cell(column=col - 1, row=row).font = Font(size=11, bold=True)
    colomn = col
    g = 0
    for val in v[i]:
        if len(v[i]) == 3:
            if val == int:
                ws.cell(column=colomn, row=row).border = left_border
            elif val == 'USD':
                ws.cell(column=colomn, row=row).border = right_border
            elif val == 'RUB':
                ws.cell(column=colomn, row=row).border = right_border
            else:
                ws.cell(column=colomn, row=row).border = thin_border
            ws.cell(column=colomn, row=row).value = val
            colomn += 1
        else:
            for d in val:
                if g in [1, 4, 7, 10]:
                    ws.column_dimensions[bukva[col + g]].width = 10
                if d == int:
                    ws.cell(column=colomn, row=row).border = left_border
                elif d == 'USD':
                    ws.cell(column=colomn, row=row).border = right_border
                elif d == 'RUB':
                    ws.cell(column=colomn, row=row).border = right_border
                else:
                    ws.cell(column=colomn, row=row).border = thin_border
                ws.cell(column=colomn, row=row).value = d
                colomn+=1
                g+=1

try:
    wb.save(name_fail)
except:
    da = input('Закройте, пожалуйста, файл '+name_fail+' и введите 1: ')
    wb.save(name_fail)




