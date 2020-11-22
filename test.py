from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from datetime import *

# определяем стили
font = Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

fill = PatternFill(fill_type='solid',
                   start_color='c1c1c1',
                   end_color='c2c2c2')

border = Border(left=Side(border_style='thin',
                          color='FF000000'),
                right=Side(border_style='thin',
                           color='FF000000'),
                top=Side(border_style='thin',
                         color='FF000000'),
                bottom=Side(border_style='thin',
                            color='FF000000'),
                diagonal=Side(border_style='thin',
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style='thin',
                             color='FF000000'),
                vertical=Side(border_style='thin',
                              color='FF000000'),
                horizontal=Side(border_style='thin',
                                color='FF000000')
                )
align_center = Alignment(horizontal='center',
                         vertical='bottom',
                         text_rotation=0,
                         wrap_text=False,
                         shrink_to_fit=False,
                         indent=0)
align_left = Alignment(horizontal='left',
                       vertical='bottom',
                       text_rotation=0,
                       wrap_text=False,
                       shrink_to_fit=False,
                       indent=0)
number_format = 'General'
protection = Protection(locked=True,
                        hidden=False)

# объект
wb = Workbook()

# активный лист
ws = wb.active

# название страницы
# ws = wb.create_sheet('первая страница', 0)
ws.title = 'первая страница'

# значение ячейки
# ws['A1'] = "Hello!"

# текущее время
today = datetime.today()
today = today.strftime('%d.%m.%Y %S:%M:%H')

# данные для строк
rows = [
    ['Название', 'Язык', 'Время'],
    ['Ivan', 'PHP', today],
    ['Egor', 'Python', today],
    ['Anton', 'Ruby', today],
    ['Roman', 'Javascript', today],
]

# циклом записываем данные
for row in rows:
    ws.append(row)

# раскрвшивание фона для заголовков
ws['A1'].fill = fill
ws['B1'].fill = fill
ws['C1'].fill = fill

# шрифты
ws['A3'].font = font
# обводка
ws['A3'].border = border
# выравнивание
ws['A3'].alignment = align_center

# вручную устанавливаем высоту первой строки
# rd = ws.row_dimensions[1]
# rd.height = 16

# увеличиваем все строки по высоте
max_row = ws.max_row
i = 1
while i <= max_row:
    rd = ws.row_dimensions[i]
    rd.height = 16
    i += 1

# сетка + выравнивание
for cellObj in ws['A1:C5']:
    for cell in cellObj:
        # print(cell.coordinate, cell.value)
        ws[cell.coordinate].border = border
        ws[cell.coordinate].alignment = align_center

# выравнивание столбца
for cellObj in ws['A2:A5']:
    for cell in cellObj:
        ws[cell.coordinate].alignment = align_left

# перетягивание ячеек
# https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    # value * коэфициент
    ws.column_dimensions[col].width = value * 1.5

# сохранение файла в текущую директорию
wb.save("sample.xlsx")

# value = pd.Series(x[0], name='Средняя цена покупки')
# currency = pd.Series(x[1], name='Валюта')
# balance = pd.Series(x[2], name='Кол-во акций')
# ticker = pd.Series(x[3], name='Тикер')
# name = pd.Series(x[4], name='Название')
# instrument_type = pd.Series(x[8], name='Тип')
# amount_all = pd.Series(x[5], name='Стоимость ин-та в п-ле')
# current_amount = pd.Series(x[6], name='Текущая стоимость ин-та')
# current_amount_all = pd.Series(x[7], name='Текущая общая стоимость')
#
# df = pd.concat([name, instrument_type, ticker, balance, value, amount_all, currency, current_amount, current_amount_all], axis=1)
# df.round(1)
# df.to_excel('example.xlsx')
# wb = load_workbook('example.xlsx')
# ws = wb.active