from openpyxl import load_workbook
import pandas as pd
from openpyxl import Workbook
from robot import get_instrument
from robot import get_operation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font
from datetime import datetime
from pytz import timezone


def list_alphabet():
    data = [i for i in range(1, 26)]
    alphabet = []
    for i in range(65, 91):
        alphabet.append(chr(i))
    la = {x: y for x,y in zip(data,alphabet)}
    return la


def set_table_1(list, col, row, title):
    ws.cell(column=col, row=row).value = title
    ws.cell(column=col, row=row).border = medium_border
    ws.cell(column=col, row=row).font = Font(size=12, bold=True)
    row+=1
    for i in list:
        ws.cell(column=col, row=row).value = i
        ws.cell(column=col, row=row).border = medium_border
        row+=1


def set_table_2(list):
    ws.column_dimensions['A'].width = 25
    set_table_1(list[4], 1, 1, 'Название')
    set_table_1(list[8], 2, 1, 'Тип')
    set_table_1(list[3], 3, 1, 'Тикер')
    set_table_1(list[2], 4, 1, 'КА')
    set_table_1(list[0], 5, 1, 'СЦП')
    set_table_1(list[6], 6, 1, 'ТС')
    set_table_1(list[7], 7, 1, 'ТОС')
    set_table_1(list[1], 8, 1, 'Валюта')
    lenght = str(len(list[4])+8)
    ws.merge_cells(str('A'+lenght+':G'+lenght))
    ws.merge_cells(str('A'+str(int(lenght)+1)+':G'+str(int(lenght)+1)))
    ws.merge_cells(str('A' + str(int(lenght) + 2) + ':G' + str(int(lenght) + 2)))
    ws.merge_cells(str('A' + str(int(lenght) + 3) + ':G' + str(int(lenght) + 3)))
    ws.merge_cells(str('A' + str(int(lenght) + 4) + ':G' + str(int(lenght) + 4)))
### Оформляется итоговая таблица
    ws.cell(column=1, row=int(lenght) - 5).value = 'Общая стоимость в ₽'
    ws.cell(column=1, row=int(lenght) - 5).border = medium_border
    ws.cell(column=1, row=int(lenght) - 5).font = Font(size=12, bold=True)
    ws.cell(column=2, row=int(lenght) - 5).value = list[9]
    ws.cell(column=2, row=int(lenght) - 5).border = medium_border
    ws.cell(column=1, row=int(lenght) - 4).value = 'Свободные ₽'
    ws.cell(column=1, row=int(lenght) - 4).border = medium_border
    ws.cell(column=1, row=int(lenght) - 4).font = Font(size=12, bold=True)
    ws.cell(column=2, row=int(lenght) - 4).value = list[10]
    ws.cell(column=2, row=int(lenght) - 4).border = medium_border
    ws.cell(column=1, row=int(lenght) - 3).value = 'Итого'
    ws.cell(column=1, row=int(lenght) - 3).border = medium_border
    ws.cell(column=1, row=int(lenght) - 3).font = Font(size=12, bold=True)
    ws.cell(column=2, row=int(lenght) - 3).value = list[9]+list[10]
    ws.cell(column=2, row=int(lenght) - 3).border = medium_border
###
    ws.cell(column=1, row=int(lenght)).value = 'КА - Колличество Акций'
    ws.cell(column=1, row=int(lenght)+1).value = 'СЦП - Средняя цена акции при покупке'
    ws.cell(column=1, row=int(lenght)+2).value = 'ТС - Текущая стоимость'
    ws.cell(column=1, row=int(lenght) + 3).value = 'ТОС - Текущая общая стоимость'


wb = Workbook()
ws = wb.create_sheet('первая страница', 0)

name_fail = (str(datetime.now(tz=timezone('Europe/Moscow')).strftime("%d-%m-%Y"))+'.xlsx')

x = get_instrument.all()
v = get_operation.dividend()

col = 11
title_row = 1
k = 1

medium_border = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))

thin_border = Border(top=Side(style='medium'),
                     bottom=Side(style='medium'))

left_border = Border(left=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))

right_border = Border(right=Side(style='medium'),
                     top=Side(style='medium'),
                     bottom=Side(style='medium'))

bukva = list_alphabet()

set_table_2(x)
ws.cell(column=col, row=title_row).value = 'price'
ws.cell(column=col, row=title_row).border = medium_border
ws.cell(column=col + 1, row=title_row).value = 'date'
ws.cell(column=col + 1, row=title_row).border = medium_border
ws.cell(column=col + 2, row=title_row).value = 'currency'
ws.cell(column=col + 2, row=title_row).border = medium_border
ws.column_dimensions[bukva[col-1]].width = 20

for i in v:
    try:
        row = (x[4].index(i))+2
    except:
        row = len(x[4])+2
        k+=1
    ws.cell(column=col - 1, row=row).value=i
    ws.cell(column=col - 1, row=row).border = medium_border
    ws.cell(column=col - 1, row=row).font = Font(size=11, bold=True)
    colomn = col
    g = 0
    for val in v[i]:
        if len(v[i]) == 3:
            if val == int:
                ws.cell(column=colomn, row=row).border = left_border
            elif val == 'USD':
                ws.cell(column=colomn, row=row).border = right_border
            elif val == 'RUB':
                ws.cell(column=colomn, row=row).border = right_border
            else:
                ws.cell(column=colomn, row=row).border = thin_border
            ws.cell(column=colomn, row=row).value = val
            colomn += 1
        else:
            for d in val:
                if g in [1, 4, 7, 10]:
                    ws.column_dimensions[bukva[col + g]].width = 10
                if d == int:
                    ws.cell(column=colomn, row=row).border = left_border
                elif d == 'USD':
                    ws.cell(column=colomn, row=row).border = right_border
                elif d == 'RUB':
                    ws.cell(column=colomn, row=row).border = right_border
                else:
                    ws.cell(column=colomn, row=row).border = thin_border
                ws.cell(column=colomn, row=row).value = d
                colomn+=1
                g+=1

try:
    wb.save(name_fail)
except:
    da = input('Закройте, пожалуйста, файл '+name_fail+' и введите 1: ')
    wb.save(name_fail)




